from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import db, User, Team, Player, Round, Bid, TiebreakerBid
from config import Config
from forms import LoginForm, RegistrationForm
from werkzeug.security import generate_password_hash
import json
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config.from_object(Config)

db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.is_admin:
            return redirect(url_for('dashboard'))
        else:
            return redirect(url_for('dashboard'))
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and user.check_password(form.password.data):
            login_user(user, remember=form.remember.data)
            return redirect(url_for('dashboard'))
        flash('Invalid username or password')
    
    return render_template('login.html', form=form)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = RegistrationForm()
    if form.validate_on_submit():
        user = User(username=form.username.data)
        user.set_password(form.password.data)
        db.session.add(user)
        
        team = Team(name=form.team_name.data, user=user)
        db.session.add(team)
        
        db.session.commit()
        flash('Your account has been created! You can now log in.', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html', form=form)

@app.route('/dashboard')
@login_required
def dashboard():
    # Check if any active rounds have expired
    active_rounds = Round.query.filter_by(is_active=True, is_tiebreaker=False).all()
    for round in active_rounds:
        if round.is_timer_expired():
            finalize_round_internal(round.id)
    
    # Check if any active tiebreaker rounds have expired
    active_tiebreakers = Round.query.filter_by(is_active=True, is_tiebreaker=True).all()
    for round in active_tiebreakers:
        if round.is_timer_expired():
            finalize_tiebreaker_round(round.id)
    
    # Refresh the data after potentially finalizing rounds
    if current_user.is_admin:
        teams = Team.query.all()
        active_rounds = Round.query.filter_by(is_active=True, is_tiebreaker=False).all()
        active_tiebreakers = Round.query.filter_by(is_active=True, is_tiebreaker=True).all()
        rounds = Round.query.all()
        return render_template('admin_dashboard.html', 
                              teams=teams, 
                              active_rounds=active_rounds,
                              active_tiebreakers=active_tiebreakers,
                              rounds=rounds, 
                              config=Config)
    
    # For team users, get only the tiebreakers they're participating in
    team_id = current_user.team.id
    tiebreaker_round_ids = db.session.query(TiebreakerBid.round_id).filter_by(team_id=team_id).distinct().all()
    tiebreaker_round_ids = [r[0] for r in tiebreaker_round_ids]
    
    active_tiebreakers = Round.query.filter(
        Round.id.in_(tiebreaker_round_ids) if tiebreaker_round_ids else False,
        Round.is_active == True,
        Round.is_tiebreaker == True
    ).all()
    
    active_rounds = Round.query.filter_by(is_active=True, is_tiebreaker=False).all()
    past_rounds = Round.query.filter_by(is_active=False).all()
    
    return render_template('team_dashboard.html', 
                          active_rounds=active_rounds,
                          active_tiebreakers=active_tiebreakers,
                          past_rounds=past_rounds, 
                          config=Config)

@app.route('/team/round')
@login_required
def team_round():
    if current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    active_rounds = Round.query.filter_by(is_active=True).all()
    
    # No active rounds, redirect to dashboard
    if not active_rounds:
        flash('There are no active rounds at the moment.', 'info')
        return redirect(url_for('dashboard'))
    
    # Get the first active round
    active_round = active_rounds[0]
    
    return render_template('team_round.html', active_round=active_round)

@app.route('/team/tiebreaker/<int:round_id>')
@login_required
def team_tiebreaker(round_id):
    """Render the tiebreaker interface for a specific tiebreaker round"""
    if current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    # Get the tiebreaker round
    tiebreaker_round = Round.query.get_or_404(round_id)
    
    # Check if it's a tiebreaker round
    if not tiebreaker_round.is_tiebreaker:
        flash('This is not a tiebreaker round.', 'error')
        return redirect(url_for('dashboard'))
    
    # Check if the team is part of this tiebreaker
    team_id = current_user.team.id
    tiebreaker_bid = TiebreakerBid.query.filter_by(
        team_id=team_id,
        round_id=round_id
    ).first()
    
    if not tiebreaker_bid:
        flash('Your team is not part of this tiebreaker.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get the player for this tiebreaker
    player = Player.query.get(tiebreaker_round.player_id)
    
    # Get all competing teams
    competing_teams = db.session.query(Team).join(
        TiebreakerBid, TiebreakerBid.team_id == Team.id
    ).filter(
        TiebreakerBid.round_id == round_id
    ).all()
    
    # Check if the round has expired
    is_expired = tiebreaker_round.is_timer_expired()
    if is_expired:
        # Finalize the round
        finalize_tiebreaker_round(round_id)
        flash('This tiebreaker round has ended.', 'info')
        return redirect(url_for('dashboard'))
    
    return render_template(
        'team_tiebreaker.html',
        tiebreaker_round=tiebreaker_round,
        player=player,
        competing_teams=competing_teams,
        current_bid=tiebreaker_bid
    )

@app.route('/team/players')
@login_required
def team_players():
    if current_user.is_admin:
        return redirect(url_for('dashboard'))
    return render_template('team_players.html')

@app.route('/team/players/data')
@login_required
def team_players_data():
    if not current_user.team:
        flash('You need to be part of a team to view this page.', 'error')
        return redirect(url_for('dashboard'))
    return render_template('team_players_data.html')

@app.route('/team/bids')
@login_required
def team_bids():
    if current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    active_rounds = Round.query.filter_by(is_active=True).all()
    return render_template('team_bids.html', active_rounds=active_rounds)

@app.route('/api/player/<int:player_id>')
@login_required
def get_player(player_id):
    player = Player.query.get_or_404(player_id)
    
    player_data = {
        'id': player.id,
        'name': player.name,
        'position': player.position,
        'overall_rating': player.overall_rating,
        'team_id': player.team_id,
        # Include existing stats
        'speed': player.speed if hasattr(player, 'speed') else None,
        'dribbling': player.dribbling if hasattr(player, 'dribbling') else None,
        'offensive_awareness': player.offensive_awareness if hasattr(player, 'offensive_awareness') else None,
        'ball_control': player.ball_control if hasattr(player, 'ball_control') else None,
        'physical_contact': player.physical_contact if hasattr(player, 'physical_contact') else None
    }
    
    return jsonify(player_data)

@app.route('/api/round/<int:round_id>')
@login_required
def get_round(round_id):
    """API endpoint to get round details"""
    round = Round.query.get_or_404(round_id)
    
    # Get all players in this round
    players_data = []
    for player in round.players:
        player_data = {
            'id': player.id,
            'name': player.name,
            'position': player.position,
            'overall_rating': player.overall_rating,
            'team': None,
            'bid_amount': None
        }
        
        # Get winning bid for this player
        if not round.is_active:
            highest_bid = Bid.query.filter_by(player_id=player.id, round_id=round_id).order_by(Bid.amount.desc()).first()
            if highest_bid:
                player_data['bid_amount'] = highest_bid.amount
                player_data['team'] = {
                    'id': highest_bid.team.id,
                    'name': highest_bid.team.name
                }
        
        # If player is assigned to a team
        if player.team_id:
            player_data['team'] = {
                'id': player.team.id,
                'name': player.team.name
            }
            
        players_data.append(player_data)
    
    round_data = {
        'id': round.id,
        'position': round.position,
        'start_time': round.start_time.isoformat() if round.start_time else None,
        'duration': round.duration,
        'is_active': round.is_active,
        'max_bids_per_team': round.max_bids_per_team,
        'players': players_data
    }
    
    return jsonify(round_data)

@app.route('/start_round', methods=['POST'])
@login_required
def start_round():
    if not current_user.is_admin:
        return jsonify({'error': 'Unauthorized'}), 403
    
    position = request.json.get('position')
    duration = request.json.get('duration', 300)  # Default to 5 minutes (300 seconds)
    max_bids_per_team = request.json.get('max_bids_per_team', 1)  # Default to 1 bid per team
    
    if not position or position not in Config.POSITIONS:
        return jsonify({'error': 'Invalid position'}), 400
    
    try:
        duration = int(duration)
        if duration < 30:  # Minimum duration of 30 seconds
            return jsonify({'error': 'Duration must be at least 30 seconds'}), 400
        # No maximum limit, allowing admin to set any reasonable duration
    except ValueError:
        return jsonify({'error': 'Invalid duration value'}), 400
    
    try:
        max_bids_per_team = int(max_bids_per_team)
        if max_bids_per_team < 1:  # Minimum of 1 bid per team
            return jsonify({'error': 'Maximum bids per team must be at least 1'}), 400
    except ValueError:
        return jsonify({'error': 'Invalid maximum bids per team value'}), 400
    
    # Check if there's already an active round
    active_round = Round.query.filter_by(is_active=True).first()
    if active_round:
        return jsonify({
            'error': f'There is already an active round for position {active_round.position}. Please finalize it before starting a new round.'
        }), 400
    
    # Create new round with timer
    round = Round(
        position=position, 
        start_time=datetime.utcnow(),
        duration=duration,
        max_bids_per_team=max_bids_per_team
    )
    db.session.add(round)
    db.session.flush()  # Get the round ID
    
    # Add players of the specified position to the round
    players = Player.query.filter_by(position=position, team_id=None).all()
    for player in players:
        player.round_id = round.id
    
    db.session.commit()
    
    return jsonify({
        'message': 'Round started successfully',
        'round_id': round.id,
        'player_count': len(players),
        'duration': duration,
        'max_bids_per_team': max_bids_per_team,
        'expires_at': (round.start_time.isoformat() if round.start_time else None)
    })

@app.route('/update_round_timer/<int:round_id>', methods=['POST'])
@login_required
def update_round_timer(round_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Unauthorized'}), 403
    
    round = Round.query.get_or_404(round_id)
    if not round.is_active:
        return jsonify({'error': 'Cannot update timer for an inactive round'}), 400
    
    duration = request.json.get('duration')
    if not duration:
        return jsonify({'error': 'Duration is required'}), 400
    
    try:
        duration = int(duration)
        if duration < 30:  # Minimum duration of 30 seconds
            return jsonify({'error': 'Duration must be at least 30 seconds'}), 400
        # No maximum limit, allowing admin to set any reasonable duration
    except ValueError:
        return jsonify({'error': 'Invalid duration value'}), 400
    
    round.duration = duration
    round.start_time = datetime.utcnow()  # Reset the timer
    db.session.commit()
    
    return jsonify({
        'message': 'Round timer updated successfully',
        'duration': duration,
        'expires_at': (round.start_time.isoformat() if round.start_time else None)
    })

@app.route('/check_round_status/<int:round_id>')
@login_required
def check_round_status(round_id):
    round = Round.query.get_or_404(round_id)
    current_time = datetime.utcnow()
    time_elapsed = (current_time - round.start_time).total_seconds()
    remaining = max(0, round.duration - time_elapsed)
    
    status = 'active'
    
    # Check if the round is already inactive
    if not round.is_active:
        status = 'ended'
    # Check if the timer has expired but round is still active
    elif time_elapsed >= round.duration:
        # Timer expired - finalize the round automatically
        if round.is_tiebreaker:
            finalize_tiebreaker_round(round.id)
        else:
            finalize_round_internal(round.id)
        status = 'ended'
    
    return jsonify({
        'status': status,
        'active': status == 'active',
        'remaining': remaining
    })

def finalize_round_internal(round_id):
    """Internal function to finalize a round, can be called programmatically"""
    round = Round.query.get(round_id)
    if not round or not round.is_active:
        return False
    
    print(f"Starting round finalization for round {round_id} (position: {round.position})")
    
    # Get all bids for this round and sort by amount (highest first)
    all_bids = Bid.query.filter_by(round_id=round_id).all()
    
    # Group bids by player and team to find the latest bid for each player-team combination
    # This handles edited bids by keeping only the latest bid amount for each team-player pair
    latest_bids = {}
    for bid in all_bids:
        key = f"{bid.player_id}-{bid.team_id}"
        if key not in latest_bids or bid.timestamp > latest_bids[key].timestamp:
            latest_bids[key] = bid
    
    # Convert to list and sort by amount (highest first)
    sorted_bids = sorted(latest_bids.values(), key=lambda x: x.amount, reverse=True)
    
    print(f"Found {len(sorted_bids)} unique bids after handling edited values")
    
    # Keep track of allocated teams and players to prevent duplicates
    allocated_teams = set()
    allocated_players = set()
    
    # Track ties that need tiebreaker rounds
    ties_to_resolve = []
    
    # First pass: identify ties and create tiebreaker rounds
    i = 0
    while i < len(sorted_bids):
        current_bid = sorted_bids[i]
        
        # Skip if team or player already allocated
        if current_bid.team_id in allocated_teams or current_bid.player_id in allocated_players:
            i += 1
            continue
            
        # Check for ties (same amount for same player)
        tied_bids = []
        j = i
        while j < len(sorted_bids) and sorted_bids[j].amount == current_bid.amount and sorted_bids[j].player_id == current_bid.player_id:
            if sorted_bids[j].team_id not in allocated_teams:
                tied_bids.append(sorted_bids[j])
            j += 1
            
        # Process ties
        if len(tied_bids) > 1:
            print(f"Found tie: {len(tied_bids)} teams bid {current_bid.amount} for player ID {current_bid.player_id}")
            # Create a tiebreaker round
            player = Player.query.get(current_bid.player_id)
            tiebreaker_round = Round(
                position=round.position,
                is_active=True,
                is_tiebreaker=True,
                parent_round_id=round.id,
                player_id=current_bid.player_id,
                duration=180  # 3 minutes for tiebreaker
            )
            db.session.add(tiebreaker_round)
            db.session.flush()  # Get the ID without committing
            
            # Add the tied teams to the tiebreaker round
            for tied_bid in tied_bids:
                tiebreaker_bid = TiebreakerBid(
                    team_id=tied_bid.team_id,
                    player_id=tied_bid.player_id,
                    round_id=tiebreaker_round.id,
                    amount=0  # Initial bid amount for tiebreaker
                )
                db.session.add(tiebreaker_bid)
                
            # Store tie information for later processing
            ties_to_resolve.append({
                'tiebreaker_round_id': tiebreaker_round.id,
                'player_id': current_bid.player_id,
                'bid_amount': current_bid.amount,
                'tied_teams': [bid.team_id for bid in tied_bids]
            })
            
            # Skip all these tied bids
            i = j
        else:
            # No tie, just a single bid
            i += 1
    
    # Commit tiebreaker rounds to the database
    db.session.commit()
    
    # If we have ties to resolve, we need to pause the main finalization
    if ties_to_resolve:
        print(f"Round {round_id} has {len(ties_to_resolve)} ties to resolve. Pausing main finalization.")
        round.is_active = False  # Deactivate the main round
        db.session.commit()
        return True
        
    # Second pass: allocate players to teams based on highest bids
    print("No ties found. Proceeding with player allocation.")
    for bid in sorted_bids:
        # Skip if team or player already allocated
        if bid.team_id in allocated_teams or bid.player_id in allocated_players:
            continue
            
        # Allocate player to team
        print(f"Allocating player ID {bid.player_id} to team ID {bid.team_id} for {bid.amount}")
        player = Player.query.get(bid.player_id)
        team = Team.query.get(bid.team_id)
        
        # Update team's balance
        team.balance -= bid.amount
        
        # Set the player's team
        player.team_id = team.id
        
        # Mark as allocated
        allocated_teams.add(team.id)
        allocated_players.add(player.id)
    
    round.is_active = False
    db.session.commit()
    print(f"Round {round_id} finalization completed. {len(allocated_teams)} allocations made.")
    return True

@app.route('/finalize_round/<int:round_id>', methods=['POST'])
@login_required
def finalize_round(round_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Unauthorized'}), 403
    
    round = Round.query.get_or_404(round_id)
    if not round.is_active:
        return jsonify({'error': 'Round already finalized'}), 400
    
    success = finalize_round_internal(round_id)
    if success:
        return jsonify({'message': 'Round finalized successfully'})
    else:
        return jsonify({'error': 'Failed to finalize round'}), 500

@app.route('/place_bid', methods=['POST'])
@login_required
def place_bid():
    if current_user.is_admin:
        return jsonify({'error': 'Admins cannot place bids'}), 403
    
    data = request.json
    round_id = data.get('round_id')
    player_id = data.get('player_id')
    amount = data.get('amount')
    
    if not all([round_id, player_id, amount]):
        return jsonify({'error': 'Missing required fields'}), 400
    
    round = Round.query.get_or_404(round_id)
    if round.is_timer_expired():
        # If timer expired, finalize the round and reject the bid
        finalize_round_internal(round_id)
        return jsonify({'error': 'Round timer has expired'}), 400
    
    if amount < Config.MINIMUM_BID:
        return jsonify({'error': f'Bid must be at least {Config.MINIMUM_BID}'}), 400
    
    team = current_user.team
    if team.balance < amount:
        return jsonify({'error': 'Insufficient balance'}), 400
    
    # Check if team has reached the maximum number of bids (20)
    total_bids = Bid.query.filter_by(team_id=team.id).count()
    if total_bids >= 20:
        return jsonify({'error': 'Maximum bid limit (20) reached for your team'}), 400
    
    # Count number of unique players this team has bid on in this round
    unique_player_bids_count = db.session.query(Bid.player_id).filter_by(
        team_id=team.id, 
        round_id=round_id
    ).distinct().count()
    
    # If bidding on a new player, check if already at the max bids per team limit
    has_bid_on_player = Bid.query.filter_by(
        team_id=team.id,
        round_id=round_id,
        player_id=player_id
    ).first() is not None
    
    if not has_bid_on_player and unique_player_bids_count >= round.max_bids_per_team:
        return jsonify({
            'error': f'You can only bid on {round.max_bids_per_team} player(s) per round. Please delete an existing bid if you wish to bid on a different player.'
        }), 400
    
    # Check if the team has already placed this bid amount on the same player
    existing_same_amount_bid = Bid.query.filter_by(
        team_id=team.id,
        round_id=round_id,
        player_id=player_id,
        amount=amount
    ).first()
    
    if existing_same_amount_bid:
        return jsonify({
            'error': f'You have already placed a bid of {amount} on this player. Please use a different amount.'
        }), 400
    
    # Check if the team has already placed this bid amount on ANY player in this round
    existing_bid_with_same_amount = Bid.query.filter_by(
        team_id=team.id,
        round_id=round_id,
        amount=amount
    ).first()
    
    if existing_bid_with_same_amount:
        other_player = Player.query.get(existing_bid_with_same_amount.player_id)
        return jsonify({
            'error': f'You have already placed a bid of {amount} on {other_player.name}. Each bid amount must be unique across all players in a round.'
        }), 400
    
    bid = Bid(
        team_id=team.id,
        player_id=player_id,
        round_id=round_id,
        amount=amount
    )
    db.session.add(bid)
    db.session.commit()
    
    return jsonify({'message': 'Bid placed successfully'})

@app.route('/delete_bid/<int:bid_id>', methods=['DELETE'])
@login_required
def delete_bid(bid_id):
    if current_user.is_admin:
        return jsonify({'error': 'Admins cannot delete bids'}), 403
    
    bid = Bid.query.get_or_404(bid_id)
    round = Round.query.get(bid.round_id)
    
    # Check if the round timer has expired
    if round and round.is_timer_expired():
        # If timer expired, finalize the round and reject the deletion
        finalize_round_internal(round.id)
        return jsonify({'error': 'Round timer has expired'}), 400
    
    # Check if the bid belongs to the current team
    if bid.team_id != current_user.team.id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    # Check if the round is still active
    if not bid.round.is_active:
        return jsonify({'error': 'Cannot delete bid from a finalized round'}), 400
    
    db.session.delete(bid)
    db.session.commit()
    
    return jsonify({'message': 'Bid deleted successfully'})

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

@app.route('/place_tiebreaker_bid', methods=['POST'])
@login_required
def place_tiebreaker_bid():
    if current_user.is_admin:
        return jsonify({'error': 'Admins cannot place bids'}), 403
    
    data = request.json
    round_id = data.get('round_id')
    player_id = data.get('player_id')
    amount = data.get('amount')
    
    if not all([round_id, player_id, amount]):
        return jsonify({'error': 'Missing required fields'}), 400
    
    round = Round.query.get_or_404(round_id)
    if not round.is_tiebreaker:
        return jsonify({'error': 'This is not a tiebreaker round'}), 400
    
    if round.is_timer_expired():
        finalize_tiebreaker_round(round_id)
        return jsonify({'error': 'Round timer has expired'}), 400
    
    team = current_user.team
    if team.balance < amount:
        return jsonify({'error': 'Insufficient balance'}), 400
    
    # Check if team is allowed in this tiebreaker
    existing_bid = TiebreakerBid.query.filter_by(
        team_id=team.id,
        round_id=round_id
    ).first()
    
    if not existing_bid:
        return jsonify({'error': 'Your team is not part of this tiebreaker'}), 403
    
    # Update the bid amount
    existing_bid.amount = amount
    db.session.commit()
    
    return jsonify({'message': 'Tiebreaker bid placed successfully'})

def finalize_tiebreaker_round(round_id):
    """Finalize a tiebreaker round and allocate the player"""
    round = Round.query.get(round_id)
    if not round or not round.is_tiebreaker or not round.is_active:
        return False
    
    print(f"Finalizing tiebreaker round {round_id} for player ID {round.player_id}")
    
    # Get all bids for this tiebreaker round
    bids = TiebreakerBid.query.filter_by(round_id=round_id).all()
    if not bids:
        print("No bids found in tiebreaker round")
        round.is_active = False
        db.session.commit()
        return False
    
    # Get the parent round
    parent_round = Round.query.get(round.parent_round_id)
    if not parent_round:
        print(f"Parent round {round.parent_round_id} not found")
        round.is_active = False
        db.session.commit()
        return False
    
    # Get teams that already have a player allocated from the parent round
    allocated_teams = set()
    allocated_players = set()
    
    # Find allocations from parent round
    for player in Player.query.filter_by(round_id=parent_round.id).all():
        if player.team_id:
            allocated_teams.add(player.team_id)
            allocated_players.add(player.id)
    
    # Find allocations from other tiebreaker rounds of the same parent
    other_tiebreakers = Round.query.filter(
        Round.parent_round_id == round.parent_round_id,
        Round.id != round.id,
        Round.is_active == False
    ).all()
    
    for tiebreaker in other_tiebreakers:
        player = Player.query.get(tiebreaker.player_id)
        if player and player.team_id:
            allocated_teams.add(player.team_id)
    
    # Filter out teams that already have a player
    eligible_bids = [bid for bid in bids if bid.team_id not in allocated_teams]
    
    # If no eligible bids left, player stays unallocated
    if not eligible_bids:
        print("No eligible bids remain after filtering allocated teams")
        round.is_active = False
        db.session.commit()
        return True
    
    # Sort eligible bids by amount (highest first)
    eligible_bids.sort(key=lambda x: x.amount, reverse=True)
    highest_bid = eligible_bids[0]
    
    # Check for another tie at the highest bid amount
    tied_bids = [bid for bid in eligible_bids if bid.amount == highest_bid.amount]
    if len(tied_bids) > 1:
        print(f"Found another tie in tiebreaker round with {len(tied_bids)} teams")
        # Create another tiebreaker round
        new_tiebreaker = Round(
            position=round.position,
            is_active=True,
            is_tiebreaker=True,
            parent_round_id=round.parent_round_id,  # Keep the original parent
            player_id=round.player_id,
            duration=180
        )
        db.session.add(new_tiebreaker)
        db.session.commit()
        
        # Add the tied teams to the new tiebreaker
        for bid in tied_bids:
            new_bid = TiebreakerBid(
                team_id=bid.team_id,
                player_id=bid.player_id,
                round_id=new_tiebreaker.id,
                amount=0
            )
            db.session.add(new_bid)
    else:
        # We have a winner - allocate the player
        print(f"Allocating player ID {round.player_id} to team ID {highest_bid.team_id} for {highest_bid.amount}")
        player = Player.query.get(round.player_id)
        team = Team.query.get(highest_bid.team_id)
        
        # For tiebreaker rounds, we need to find the original bid amount from the parent round
        original_bid = Bid.query.filter_by(
            round_id=round.parent_round_id,
            team_id=highest_bid.team_id,
            player_id=round.player_id
        ).first()
        
        if original_bid:
            team.balance -= original_bid.amount
            player.team_id = team.id
            
            # Mark this player and team as allocated
            allocated_teams.add(team.id)
            allocated_players.add(player.id)
        else:
            print(f"Warning: Could not find original bid for team {team.id} on player {player.id}")
    
    # Deactivate this tiebreaker round
    round.is_active = False
    db.session.commit()
    
    # Check if the parent round has any unresolved tiebreakers
    active_tiebreakers = Round.query.filter_by(
        parent_round_id=round.parent_round_id,
        is_active=True
    ).count()
    
    # If parent round has no active tiebreakers, resume its finalization
    if active_tiebreakers == 0 and not parent_round.is_active:
        print(f"All tiebreakers resolved for parent round {parent_round.id}, resuming finalization")
        # Continue with the auction process
        resume_after_tiebreaker(parent_round.id)
    
    return True

def resume_after_tiebreaker(round_id):
    """Resume the finalization process after tiebreakers have been resolved"""
    round = Round.query.get(round_id)
    if not round:
        return False
    
    print(f"Resuming finalization for round {round_id} after tiebreakers")
    
    # Get latest bids (edited bids handled)
    all_bids = Bid.query.filter_by(round_id=round_id).all()
    latest_bids = {}
    for bid in all_bids:
        key = f"{bid.player_id}-{bid.team_id}"
        if key not in latest_bids or bid.timestamp > latest_bids[key].timestamp:
            latest_bids[key] = bid
    
    # Convert to list and sort by amount (highest first)
    sorted_bids = sorted(latest_bids.values(), key=lambda x: x.amount, reverse=True)
    
    # Find currently allocated teams and players
    allocated_teams = set()
    allocated_players = set()
    
    for player in Player.query.filter_by(round_id=round_id).all():
        if player.team_id:
            allocated_teams.add(player.team_id)
            allocated_players.add(player.id)
    
    # Also check tiebreaker rounds
    for tiebreaker in Round.query.filter_by(parent_round_id=round_id).all():
        if tiebreaker.player_id and not tiebreaker.is_active:
            player = Player.query.get(tiebreaker.player_id)
            if player and player.team_id:
                allocated_teams.add(player.team_id)
                allocated_players.add(player.id)
    
    # Process remaining bids
    for bid in sorted_bids:
        # Skip if team or player already allocated
        if bid.team_id in allocated_teams or bid.player_id in allocated_players:
            continue
            
        # No need to check for ties - they should have been handled already
        # Allocate player to team
        print(f"Allocating player ID {bid.player_id} to team ID {bid.team_id} for {bid.amount}")
        player = Player.query.get(bid.player_id)
        team = Team.query.get(bid.team_id)
        
        # Update team's balance
        team.balance -= bid.amount
        
        # Set the player's team
        player.team_id = team.id
        
        # Mark as allocated
        allocated_teams.add(team.id)
        allocated_players.add(player.id)
    
    db.session.commit()
    print(f"Finalization completed for round {round_id} after tiebreakers")
    return True

@app.route('/finalize_tiebreaker/<int:round_id>', methods=['POST'])
@login_required
def finalize_tiebreaker(round_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Only admins can finalize rounds'}), 403
    
    if finalize_tiebreaker_round(round_id):
        return jsonify({'message': 'Tiebreaker round finalized successfully'})
    else:
        return jsonify({'error': 'Failed to finalize tiebreaker round'}), 400

@app.route('/admin/teams')
@login_required
def admin_teams():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    teams = Team.query.all()
    team_details = []
    
    for team in teams:
        # Get position counts
        position_counts = {}
        for position in Config.POSITIONS:
            position_counts[position] = 0
        
        # Calculate total team value and get position distribution
        total_team_value = 0
        players_by_position = {}
        
        for player in team.players:
            position_counts[player.position] += 1
            
            # Get player's acquisition value (from winning bid)
            acquisition_value = 0
            winning_bid = Bid.query.filter_by(
                team_id=team.id,
                player_id=player.id
            ).order_by(Bid.amount.desc()).first()
            
            if winning_bid:
                acquisition_value = winning_bid.amount
            
            total_team_value += acquisition_value
            
            # Group players by position
            if player.position not in players_by_position:
                players_by_position[player.position] = []
            
            players_by_position[player.position].append({
                'id': player.id,
                'name': player.name,
                'overall_rating': player.overall_rating,
                'acquisition_value': acquisition_value
            })
        
        # Get active and completed bid counts
        active_bids_count = 0
        completed_rounds = Round.query.filter_by(is_active=False).all()
        completed_round_ids = [r.id for r in completed_rounds]
        
        # Get active rounds
        active_rounds = Round.query.filter_by(is_active=True).all()
        
        # Count active bids
        for round in active_rounds:
            active_bids_count += Bid.query.filter_by(team_id=team.id, round_id=round.id).count()
        
        # Count completed bids
        completed_bids_count = Bid.query.filter(
            Bid.team_id == team.id,
            Bid.round_id.in_(completed_round_ids)
        ).count()
        
        # Find if team has a user
        team_user = User.query.filter_by(id=team.user_id).first()
        
        team_details.append({
            'team': team,
            'position_counts': position_counts,
            'total_players': len(team.players),
            'total_team_value': total_team_value,
            'active_bids_count': active_bids_count,
            'completed_bids_count': completed_bids_count,
            'players_by_position': players_by_position,
            'has_user': team_user is not None,
            'username': team_user.username if team_user else None
        })
    
    return render_template('admin_teams.html', teams=team_details)

@app.route('/admin/players')
@login_required
def admin_players():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    # Get query parameters for pagination and filtering
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)  # Default to 20 players per page
    position_filter = request.args.get('position', None)
    search_query = request.args.get('q', None)
    
    # Build the query
    query = Player.query
    
    # Apply position filter if provided
    if position_filter and position_filter in Config.POSITIONS:
        query = query.filter_by(position=position_filter)
    
    # Apply search filter if provided
    if search_query and search_query.strip():
        search_term = f"%{search_query.strip()}%"
        query = query.filter(Player.name.ilike(search_term))
    
    # Get paginated results
    players_pagination = query.order_by(Player.overall_rating.desc()).paginate(page=page, per_page=per_page)
    
    # Get all teams for the edit form
    teams = Team.query.all()
    
    return render_template(
        'admin_players.html', 
        players_pagination=players_pagination,
        players=players_pagination.items,
        teams=teams,
        config=Config,
        current_position=position_filter,
        search_query=search_query
    )

@app.route('/admin/rounds')
@login_required
def admin_rounds():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    # Get active rounds
    active_rounds = Round.query.filter_by(is_active=True).all()
    
    # Get past (inactive) rounds
    past_rounds = Round.query.filter_by(is_active=False).all()
    
    # Get all rounds for completeness
    rounds = Round.query.all()
    
    return render_template('admin_rounds.html', 
                          rounds=rounds, 
                          active_rounds=active_rounds, 
                          past_rounds=past_rounds,
                          config=Config)

@app.route('/admin/edit_round/<int:round_id>', methods=['POST'])
@login_required
def admin_edit_round(round_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Unauthorized'}), 403
    
    round = Round.query.get_or_404(round_id)
    if not round.is_active:
        return jsonify({'error': 'Cannot edit a completed round'}), 400
    
    position = request.json.get('position')
    duration = request.json.get('duration')
    
    if position and position not in Config.POSITIONS:
        return jsonify({'error': 'Invalid position'}), 400
    
    if duration:
        try:
            duration = int(duration)
            if duration < 30:
                return jsonify({'error': 'Duration must be at least 30 seconds'}), 400
        except ValueError:
            return jsonify({'error': 'Invalid duration value'}), 400
    
    if position:
        round.position = position
    if duration:
        round.duration = duration
        round.start_time = datetime.utcnow()  # Reset the timer
    
    db.session.commit()
    return jsonify({'message': 'Round updated successfully'})

@app.route('/admin/delete_round/<int:round_id>', methods=['POST'])
@login_required
def admin_delete_round(round_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Unauthorized'}), 403
    
    round = Round.query.get_or_404(round_id)
    if round.is_active:
        return jsonify({'error': 'Cannot delete an active round'}), 400
    
    try:
        print(f"Processing round deletion - ID: {round_id}, Position: {round.position}")
        
        # Find all players from this round that have been allocated to teams
        allocated_players = []
        refund_total = 0
        freed_players_count = 0
        
        for player in round.players:
            if player.team_id is not None:
                print(f"Found allocated player: {player.name} (ID: {player.id}) to team ID: {player.team_id}")
                # Find the winning bid for this player
                winning_bid = Bid.query.filter_by(
                    round_id=round_id,
                    player_id=player.id,
                    team_id=player.team_id
                ).first()
                
                if winning_bid:
                    # Refund the team
                    team = Team.query.get(player.team_id)
                    if team:
                        team.balance += winning_bid.amount
                        refund_total += winning_bid.amount
                        print(f"Refunded {winning_bid.amount} to team: {team.name}")
                
                # Free the player
                player.team_id = None
                freed_players_count += 1
        
        # Delete all bids associated with this round
        bid_count = Bid.query.filter_by(round_id=round_id).count()
        Bid.query.filter_by(round_id=round_id).delete()
        print(f"Deleted {bid_count} bids from round")
        
        # Also delete any tiebreaker bids if this was a tiebreaker round
        tiebreaker_bid_count = 0
        if round.is_tiebreaker:
            tiebreaker_bid_count = TiebreakerBid.query.filter_by(round_id=round_id).count()
            TiebreakerBid.query.filter_by(round_id=round_id).delete()
            print(f"Deleted {tiebreaker_bid_count} tiebreaker bids from round")
            
        # Clear round_id from any players in this round
        player_count = 0
        for player in round.players:
            player.round_id = None
            player_count += 1
        print(f"Cleared round_id from {player_count} players")
            
        # Delete tiebreaker rounds that might reference this round as parent
        tiebreaker_rounds = Round.query.filter_by(parent_round_id=round_id).all()
        tiebreaker_round_count = len(tiebreaker_rounds)
        for tiebreaker in tiebreaker_rounds:
            # Recursively handle tiebreaker rounds
            tiebreaker_bids_deleted = TiebreakerBid.query.filter_by(round_id=tiebreaker.id).count()
            TiebreakerBid.query.filter_by(round_id=tiebreaker.id).delete()
            print(f"Deleted tiebreaker round ID: {tiebreaker.id} with {tiebreaker_bids_deleted} bids")
            db.session.delete(tiebreaker)
            
        # Now we can safely delete the round
        db.session.delete(round)
        db.session.commit()
        
        summary = f"Round deleted successfully. Freed {freed_players_count} players, refunded {refund_total} to teams, removed {bid_count} bids, and {tiebreaker_round_count} tiebreaker rounds."
        print(summary)
        return jsonify({'message': summary})
    except Exception as e:
        db.session.rollback()
        error_message = f"Failed to delete round: {str(e)}"
        print(f"ERROR: {error_message}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': error_message}), 500

@app.route('/admin/add_team', methods=['POST'])
@login_required
def admin_add_team():
    if not current_user.is_admin:
        return jsonify({'error': 'Unauthorized'}), 403
    
    name = request.json.get('name')
    balance = request.json.get('balance', 1000000)  # Default balance
    
    if not name:
        return jsonify({'error': 'Team name is required'}), 400
    
    if Team.query.filter_by(name=name).first():
        return jsonify({'error': 'Team name already exists'}), 400
    
    try:
        balance = int(balance)
        if balance < 0:
            return jsonify({'error': 'Balance cannot be negative'}), 400
    except ValueError:
        return jsonify({'error': 'Invalid balance value'}), 400
    
    team = Team(name=name, balance=balance)
    db.session.add(team)
    db.session.commit()
    
    return jsonify({'message': 'Team added successfully'})

@app.route('/admin/edit_team/<int:team_id>', methods=['POST'])
@login_required
def admin_edit_team(team_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Unauthorized'}), 403
    
    team = Team.query.get_or_404(team_id)
    name = request.json.get('name')
    balance = request.json.get('balance')
    
    if name and name != team.name:
        if Team.query.filter_by(name=name).first():
            return jsonify({'error': 'Team name already exists'}), 400
        team.name = name
    
    if balance is not None:
        try:
            balance = int(balance)
            if balance < 0:
                return jsonify({'error': 'Balance cannot be negative'}), 400
            team.balance = balance
        except ValueError:
            return jsonify({'error': 'Invalid balance value'}), 400
    
    db.session.commit()
    return jsonify({'message': 'Team updated successfully'})

@app.route('/admin/delete_team/<int:team_id>', methods=['POST'])
@login_required
def admin_delete_team(team_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Unauthorized'}), 403
    
    team = Team.query.get_or_404(team_id)
    
    # Check if team has any players
    if team.players:
        return jsonify({'error': 'Cannot delete a team with players'}), 400
    
    db.session.delete(team)
    db.session.commit()
    return jsonify({'message': 'Team deleted successfully'})

@app.route('/admin/add_player', methods=['POST'])
@login_required
def admin_add_player():
    if not current_user.is_admin:
        return jsonify({'error': 'Unauthorized'}), 403
    
    data = request.json
    required_fields = ['name', 'position', 'overall_rating']
    
    for field in required_fields:
        if field not in data:
            return jsonify({'error': f'{field} is required'}), 400
    
    if data['position'] not in Config.POSITIONS:
        return jsonify({'error': 'Invalid position'}), 400
    
    try:
        overall_rating = int(data['overall_rating'])
        if not (1 <= overall_rating <= 99):
            return jsonify({'error': 'Overall rating must be between 1 and 99'}), 400
    except ValueError:
        return jsonify({'error': 'Invalid overall rating value'}), 400
    
    player = Player(
        name=data['name'],
        position=data['position'],
        overall_rating=overall_rating,
        offensive_awareness=data.get('offensive_awareness', 50),
        ball_control=data.get('ball_control', 50),
        dribbling=data.get('dribbling', 50),
        tight_possession=data.get('tight_possession', 50),
        low_pass=data.get('low_pass', 50),
        lofted_pass=data.get('lofted_pass', 50),
        finishing=data.get('finishing', 50),
        heading=data.get('heading', 50),
        set_piece_taking=data.get('set_piece_taking', 50),
        curl=data.get('curl', 50),
        speed=data.get('speed', 50),
        acceleration=data.get('acceleration', 50),
        kicking_power=data.get('kicking_power', 50),
        jumping=data.get('jumping', 50),
        physical_contact=data.get('physical_contact', 50),
        balance=data.get('balance', 50),
        stamina=data.get('stamina', 50),
        defensive_awareness=data.get('defensive_awareness', 50),
        tackling=data.get('tackling', 50),
        aggression=data.get('aggression', 50),
        defensive_engagement=data.get('defensive_engagement', 50)
    )
    
    db.session.add(player)
    db.session.commit()
    
    return jsonify({'message': 'Player added successfully'})

@app.route('/admin/edit_player/<int:player_id>', methods=['POST'])
@login_required
def admin_edit_player(player_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Unauthorized'}), 403
    
    player = Player.query.get_or_404(player_id)
    data = request.json
    
    if 'position' in data and data['position'] not in Config.POSITIONS:
        return jsonify({'error': 'Invalid position'}), 400
    
    if 'overall_rating' in data:
        try:
            overall_rating = int(data['overall_rating'])
            if not (1 <= overall_rating <= 99):
                return jsonify({'error': 'Overall rating must be between 1 and 99'}), 400
            player.overall_rating = overall_rating
        except ValueError:
            return jsonify({'error': 'Invalid overall rating value'}), 400
    
    # Update other attributes if provided
    for attr in ['name', 'position', 'offensive_awareness', 'ball_control', 'dribbling',
                'tight_possession', 'low_pass', 'lofted_pass', 'finishing', 'heading',
                'set_piece_taking', 'curl', 'speed', 'acceleration', 'kicking_power',
                'jumping', 'physical_contact', 'balance', 'stamina', 'defensive_awareness',
                'tackling', 'aggression', 'defensive_engagement']:
        if attr in data:
            setattr(player, attr, data[attr])
    
    db.session.commit()
    return jsonify({'message': 'Player updated successfully'})

@app.route('/admin/delete_player/<int:player_id>', methods=['POST'])
@login_required
def admin_delete_player(player_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Unauthorized'}), 403
    
    player = Player.query.get_or_404(player_id)
    
    # Check if player is in a team
    if player.team_id:
        return jsonify({'error': 'Cannot delete a player that is in a team'}), 400
    
    # Check if player is in an active round
    if player.round_id and Round.query.get(player.round_id).is_active:
        return jsonify({'error': 'Cannot delete a player that is in an active round'}), 400
    
    db.session.delete(player)
    db.session.commit()
    return jsonify({'message': 'Player deleted successfully'})

@app.route('/admin/export_players')
@login_required
def export_players():
    """Export all players to an Excel file with sheets organized by position"""
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    try:
        import pandas as pd
        from io import BytesIO
        from openpyxl.utils import get_column_letter
        
        # Create a BytesIO object to save the Excel file
        output = BytesIO()
        
        # Create Excel writer
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Loop through each position
            for position in Config.POSITIONS:
                # Get players for this position
                players = Player.query.filter_by(position=position).all()
                
                if not players:
                    continue  # Skip if no players for this position
                
                # Create a data list for pandas DataFrame
                data = []
                for player in players:
                    # Get team name if player is in a team
                    team_name = None
                    if player.team_id:
                        team = Team.query.get(player.team_id)
                        team_name = team.name if team else None
                    
                    # Create player data dictionary
                    player_data = {
                        'ID': player.id,
                        'Name': player.name,
                        'Position': player.position,
                        'Overall Rating': player.overall_rating,
                        'Playing Style': player.playing_style if hasattr(player, 'playing_style') else None,
                        'Team': team_name,
                    }
                    
                    # Add all available attribute fields
                    all_attributes = [
                        'speed', 'dribbling', 'offensive_awareness', 'ball_control', 
                        'physical_contact', 'tight_possession', 'low_pass', 'lofted_pass',
                        'finishing', 'heading', 'set_piece_taking', 'curl', 'acceleration',
                        'kicking_power', 'jumping', 'balance', 'stamina', 'defensive_awareness',
                        'tackling', 'aggression', 'defensive_engagement'
                    ]
                    
                    for attr in all_attributes:
                        if hasattr(player, attr):
                            # Convert from snake_case to Title Case for display
                            attr_display = ' '.join(word.capitalize() for word in attr.split('_'))
                            player_data[attr_display] = getattr(player, attr)
                    
                    data.append(player_data)
                
                # Create DataFrame and write to Excel
                df = pd.DataFrame(data)
                df.to_excel(writer, sheet_name=position, index=False)
                
                # Auto-adjust columns' width
                worksheet = writer.sheets[position]
                
                # Convert all values to strings before calculating column width
                str_df = df.astype(str)
                
                for i, col in enumerate(df.columns):
                    # Convert numeric column index to Excel-style letter (A, B, C, etc.)
                    column_letter = get_column_letter(i + 1)
                    
                    # Find the maximum length in the column
                    # Handle None values by converting to empty string
                    max_len = max(
                        max([len(str(x)) for x in str_df[col] if str(x) != 'None']+[0]),  # Length of largest item
                        len(str(col))  # Length of column name
                    ) + 2  # Add a little extra space
                    
                    # Set the column width
                    worksheet.column_dimensions[column_letter].width = max_len
        
        # Seek to beginning of file
        output.seek(0)
        
        # Create response with Excel file
        from flask import send_file
        filename = f"players_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        import traceback
        print(f"Error exporting players: {str(e)}")
        traceback.print_exc()
        flash(f"Error exporting players: {str(e)}", "error")
        return redirect(url_for('admin_players'))

@app.route('/api/active_tiebreakers')
@login_required
def get_active_tiebreakers():
    """API endpoint to get active tiebreaker rounds for the current team"""
    if current_user.is_admin:
        # For admins, get all active tiebreaker rounds
        tiebreaker_rounds = Round.query.filter_by(is_active=True, is_tiebreaker=True).all()
    else:
        # For team users, get only tiebreaker rounds they're participating in
        team_id = current_user.team.id
        # Query for rounds where this team has a tiebreaker bid
        tiebreaker_rounds_ids = db.session.query(TiebreakerBid.round_id).filter_by(team_id=team_id).distinct().all()
        tiebreaker_rounds_ids = [r[0] for r in tiebreaker_rounds_ids]  # Extract IDs from result tuples
        
        # Get the active tiebreaker rounds
        tiebreaker_rounds = Round.query.filter(
            Round.id.in_(tiebreaker_rounds_ids),
            Round.is_active == True,
            Round.is_tiebreaker == True
        ).all()

    result = []
    for round in tiebreaker_rounds:
        # Get the player this tiebreaker is for
        player = Player.query.get(round.player_id)
        
        # Get all bids for this tiebreaker
        tiebreaker_bids = TiebreakerBid.query.filter_by(round_id=round.id).all()
        
        # Format bids data
        bids_data = []
        for bid in tiebreaker_bids:
            team = Team.query.get(bid.team_id)
            bids_data.append({
                'id': bid.id,
                'team_id': bid.team_id,
                'team_name': team.name,
                'amount': bid.amount,
                'is_current_team': bid.team_id == current_user.team.id if not current_user.is_admin else False
            })
        
        # Calculate time remaining
        current_time = datetime.utcnow()
        elapsed = (current_time - round.start_time).total_seconds()
        time_remaining = max(0, round.duration - elapsed)
        
        round_data = {
            'id': round.id,
            'player': {
                'id': player.id,
                'name': player.name,
                'position': player.position,
                'overall_rating': player.overall_rating
            },
            'start_time': round.start_time.isoformat(),
            'duration': round.duration,
            'time_remaining': time_remaining,
            'bids': bids_data,
            'parent_round_id': round.parent_round_id
        }
        result.append(round_data)
    
    return jsonify(result)

@app.route('/admin/round/<int:round_id>')
@login_required
def admin_view_round(round_id):
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    round = Round.query.get_or_404(round_id)
    teams = Team.query.all()
    
    # Get all bids for this round
    all_bids = Bid.query.filter_by(round_id=round_id).all()
    
    # Get team bids data
    team_bids = {}
    for team in teams:
        team_bids[team.id] = {
            'team': team,
            'bids': []
        }
    
    # Collect winning bids data
    winning_bids = []
    
    # Process all bids
    for bid in all_bids:
        # Add to team bids
        if bid.team_id in team_bids:
            player = Player.query.get(bid.player_id)
            is_winning = player.team_id == bid.team_id
            
            team_bids[bid.team_id]['bids'].append({
                'bid': bid,
                'player': player,
                'is_winning': is_winning
            })
            
            # If this is a winning bid, add to winning bids list
            if is_winning:
                winning_bids.append({
                    'bid': bid,
                    'player': player,
                    'team': team_bids[bid.team_id]['team']
                })
    
    # Sort winning bids by bid amount (descending)
    winning_bids.sort(key=lambda x: x['bid'].amount, reverse=True)
    
    return render_template(
        'admin_view_round.html',
        round=round,
        team_bids=team_bids,
        winning_bids=winning_bids,
        teams=teams
    )

@app.route('/admin/export_winning_bids')
@login_required
def export_winning_bids():
    """Export winning bids to an Excel file"""
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    try:
        import pandas as pd
        from io import BytesIO
        from openpyxl.utils import get_column_letter
        
        # Create a BytesIO object to save the Excel file
        output = BytesIO()
        
        # Get all completed rounds
        rounds = Round.query.filter_by(is_active=False).all()
        
        # Create Excel writer
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Create a data list for pandas DataFrame
            data = []
            
            for round_obj in rounds:
                # Get all players from this round that have been allocated to teams
                allocated_players = Player.query.filter(
                    (Player.round_id == round_obj.id) & 
                    (Player.team_id != None)
                ).all()
                
                # For each allocated player, find the winning bid
                for player in allocated_players:
                    # Find the winning bid (the bid from the team that won the player)
                    winning_bid = Bid.query.filter_by(
                        round_id=round_obj.id,
                        player_id=player.id,
                        team_id=player.team_id
                    ).first()
                    
                    if winning_bid:
                        team = Team.query.get(player.team_id)
                        
                        # Create bid data dictionary
                        bid_data = {
                            'Round ID': round_obj.id,
                            'Position': round_obj.position,
                            'Player Name': player.name,
                            'Player Rating': player.overall_rating,
                            'Playing Style': player.playing_style if hasattr(player, 'playing_style') else None,
                            'Team Name': team.name if team else "Unknown",
                            'Bid Amount': winning_bid.amount,
                            'Bid Date': winning_bid.timestamp.strftime('%Y-%m-%d %H:%M:%S')
                        }
                        data.append(bid_data)
            
            # Create DataFrame and write to Excel
            if data:
                df = pd.DataFrame(data)
                df.to_excel(writer, sheet_name='Winning Bids', index=False)
                
                # Auto-adjust columns' width
                worksheet = writer.sheets['Winning Bids']
                for i, col in enumerate(df.columns):
                    # Convert numeric column index to Excel-style letter
                    column_letter = get_column_letter(i + 1)
                    
                    # Find the maximum length in the column
                    # Handle None values by converting to empty string
                    max_len = max(
                        max([len(str(x)) for x in df[col] if str(x) != 'None']+[0]),  # Length of largest item
                        len(str(col))  # Length of column name
                    ) + 2  # Add a little extra space
                    
                    # Set the column width
                    worksheet.column_dimensions[column_letter].width = max_len
            else:
                # Create an empty sheet if no data
                pd.DataFrame().to_excel(writer, sheet_name='Winning Bids', index=False)
                writer.sheets['Winning Bids'].cell(row=1, column=1).value = "No winning bids found"
        
        # Seek to beginning of file
        output.seek(0)
        
        # Create response with Excel file
        from flask import send_file
        filename = f"winning_bids_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        import traceback
        print(f"Error exporting winning bids: {str(e)}")
        traceback.print_exc()
        flash(f"Error exporting winning bids: {str(e)}", "error")
        return redirect(url_for('admin_rounds'))

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000) 